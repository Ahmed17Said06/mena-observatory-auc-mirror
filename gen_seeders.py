import openpyxl, datetime

def fmt_date(val):
    if val is None:
        return 'null'
    if isinstance(val, datetime.datetime):
        return "'" + val.strftime('%Y-%m-%d') + "'"
    if isinstance(val, (int, float)):
        return "'" + str(int(val)) + "-01-01'"
    s = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
        try:
            return "'" + datetime.datetime.strptime(s, fmt).strftime('%Y-%m-%d') + "'"
        except Exception:
            pass
    if s.isdigit():
        return "'" + s + "-01-01'"
    return "'" + s + "'"

def php_str(s):
    if s is None:
        return "''"
    return "'" + str(s).replace('\\', '\\\\').replace("'", "\\'") + "'"

def tags_arr(tag_str):
    if not tag_str:
        return []
    return [t.strip() for t in str(tag_str).split(',') if t.strip()]

def repo_entry(title, inst, pub_date, link, tags, is_global):
    tp = '[' + ', '.join(php_str(t) for t in tags) + ']'
    is_g = 'true' if is_global else 'false'
    link_val = php_str(link) if link else "''"
    lines = [
        "            [",
        f"                'title'        => {php_str(title)},",
        f"                'description'  => {php_str(inst)},",
        f"                'publish_date' => {fmt_date(pub_date)},",
        f"                'data_link'    => {link_val},",
        f"                'is_global'    => {is_g},",
        "                'is_our_work'  => false,",
        "                'image'        => '',",
        f"                'tags'         => {tp},",
        "            ],",
    ]
    return '\n'.join(lines)

# ── SEEDER 1: GlobalResearchOutputsSeeder ─────────────────────────────────────
wb1 = openpyxl.load_workbook('Global Research Outputs.xlsm', keep_vba=True, data_only=True)
ws  = wb1['Global Research Outputs']

entries = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    _, title, inst, pub_date, link, tag = row
    entries.append(repo_entry(title, inst, pub_date, link, tags_arr(tag), True))

body1 = '\n'.join(entries)

seeder1 = f"""<?php

namespace Database\\Seeders;

use App\\Models\\Repo;
use App\\Models\\Repo_tags;
use Illuminate\\Database\\Seeder;

class GlobalResearchOutputsSeeder extends Seeder
{{
    public function run(): void
    {{
        ${'$'}publications = [
{body1}
        ];

        foreach (${'$'}publications as ${'$'}data) {{
            ${'$'}tagNames = ${'$'}data['tags'];
            unset(${'$'}data['tags']);

            if (Repo::where('title', ${'$'}data['title'])->exists()) {{
                continue;
            }}

            ${'$'}repo = Repo::create(${'$'}data);

            foreach (${'$'}tagNames as ${'$'}name) {{
                ${'$'}tag = Repo_tags::firstOrCreate(['name' => ${'$'}name]);
                ${'$'}repo->tags()->attach(${'$'}tag->id);
            }}
        }}
    }}
}}
"""

# Write without f-string tricks - build the PHP manually
seeder1_php = seeder1.replace("${'$'}", '$')

with open('database/seeders/GlobalResearchOutputsSeeder.php', 'w', encoding='utf-8') as f:
    f.write(seeder1_php)
print(f'Written: GlobalResearchOutputsSeeder.php  ({len(entries)} entries)')

# ── SEEDER 2: RegionalResearchAndNewsSeeder ────────────────────────────────────
wb2 = openpyxl.load_workbook('Research and News.xlsx', data_only=True)

repo_entries = []

ws_reg = wb2['Regional']
for row in ws_reg.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    _, title, inst, pub_date, link, tag = row
    repo_entries.append(repo_entry(title, inst, pub_date, link, tags_arr(tag), False))

ws_glob = wb2['Global']
for row in ws_glob.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    _, title, inst, pub_date, link, tag = row
    repo_entries.append(repo_entry(title, inst, pub_date, link, tags_arr(tag), True))

ws_news = wb2['News']
news_entries = []
for row in ws_news.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    title, source, pub_date, link = row
    desc = ('Source: ' + str(source)) if source else ''
    ne = '\n'.join([
        "            [",
        f"                'title'       => {php_str(title)},",
        f"                'date'        => {fmt_date(pub_date)},",
        f"                'data_link'   => {php_str(link)},",
        f"                'description' => {php_str(desc)},",
        "                'image'       => '',",
        "            ],",
    ])
    news_entries.append(ne)

repo_body  = '\n'.join(repo_entries)
news_body  = '\n'.join(news_entries)

seeder2_lines = [
    "<?php",
    "",
    "namespace Database\\Seeders;",
    "",
    "use App\\Models\\News;",
    "use App\\Models\\Repo;",
    "use App\\Models\\Repo_tags;",
    "use Illuminate\\Database\\Seeder;",
    "",
    "class RegionalResearchAndNewsSeeder extends Seeder",
    "{",
    "    public function run(): void",
    "    {",
    "        // -- Repos (regional and global external research) -----------------",
    "        $publications = [",
    repo_body,
    "        ];",
    "",
    "        foreach ($publications as $data) {",
    "            $tagNames = $data['tags'];",
    "            unset($data['tags']);",
    "",
    "            if (Repo::where('title', $data['title'])->exists()) {",
    "                continue;",
    "            }",
    "",
    "            $repo = Repo::create($data);",
    "",
    "            foreach ($tagNames as $name) {",
    "                $tag = Repo_tags::firstOrCreate(['name' => $name]);",
    "                $repo->tags()->attach($tag->id);",
    "            }",
    "        }",
    "",
    "        // -- News items ----------------------------------------------------",
    "        $newsItems = [",
    news_body,
    "        ];",
    "",
    "        foreach ($newsItems as $data) {",
    "            if (News::where('title', $data['title'])->exists()) {",
    "                continue;",
    "            }",
    "            News::create($data);",
    "        }",
    "    }",
    "}",
    "",
]

seeder2_php = '\n'.join(seeder2_lines)

with open('database/seeders/RegionalResearchAndNewsSeeder.php', 'w', encoding='utf-8') as f:
    f.write(seeder2_php)

print(f'Written: RegionalResearchAndNewsSeeder.php  ({len(repo_entries)} repos, {len(news_entries)} news)')
